#!/usr/bin/env python3
"""Build the daily fulfillment Excel sheet from a Shopify order payload.

The fulfillment partner must never see what we sell for. This script is
price-blind by construction, two ways:

  1. It only ever emits COLUMNS below - an allowlist, not a denylist.
  2. It REFUSES to run if the payload carries any money-bearing field
     (MONEY_KEY_RE). So if the routine's GraphQL query ever grows a price
     field, this fails loudly instead of quietly leaking it downstream.

Usage:
    python build_fulfillment_xlsx.py --orders orders.json --out "...\\order-7.17.xlsx"

--orders accepts either a raw Shopify GraphQL response
({"data": {"orders": {"edges": [{"node": {...}}]}}}) or a plain
{"orders": [ ...order nodes... ]}.

If --out already exists (a second run on the same day), rows are APPENDED to
the existing sheet rather than overwriting it.
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Order", "Order Date", "Customer", "Email", "Phone",
    "Address 1", "Address 2", "City", "State/Province", "ZIP", "Country",
    "Product", "Variant", "SKU", "Qty", "Image Link", "Note",
]

# Any payload key matching this is a money field and must never reach the sheet.
MONEY_KEY_RE = re.compile(
    r"price|amount|money|total|subtotal|discount|tax|currency|cost",
    re.IGNORECASE,
)

COL_WIDTHS = {
    "Order": 10, "Order Date": 12, "Customer": 22, "Email": 28, "Phone": 16,
    "Address 1": 30, "Address 2": 16, "City": 18, "State/Province": 14,
    "ZIP": 10, "Country": 14, "Product": 34, "Variant": 20, "SKU": 18,
    "Qty": 5, "Image Link": 46, "Note": 30,
}

HEADER_FILL = PatternFill("solid", fgColor="1F2933")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
STRIPE_FILL = PatternFill("solid", fgColor="F2F4F6")
LINK_FONT = Font(color="0563C1", underline="single")


def assert_price_blind(node, path="orders"):
    """Recursively refuse any money-bearing key. Fail loud, never leak."""
    if isinstance(node, dict):
        for key, value in node.items():
            if MONEY_KEY_RE.search(key):
                sys.exit(
                    f"REFUSING TO BUILD: money field '{key}' found at {path}.\n"
                    "The fulfillment sheet must carry no selling-price data. "
                    "Remove that field from the GraphQL query and re-run."
                )
            assert_price_blind(value, f"{path}.{key}")
    elif isinstance(node, list):
        for i, item in enumerate(node):
            assert_price_blind(item, f"{path}[{i}]")


def load_orders(path):
    # utf-8-sig: PowerShell's Out-File writes a BOM that plain utf-8 chokes on.
    with open(path, encoding="utf-8-sig") as fh:
        payload = json.load(fh)

    # Accept the raw GraphQL envelope or a plain list.
    container = payload
    for key in ("data", "orders"):
        if isinstance(container, dict) and key in container:
            container = container[key]

    if isinstance(container, dict) and "edges" in container:
        orders = [edge.get("node", {}) for edge in container["edges"]]
    elif isinstance(container, list):
        orders = container
    else:
        sys.exit(f"Could not find an order list in {path}.")

    assert_price_blind(orders)
    return orders


def unwrap(node, key):
    """Return a list from either {key: {edges: [{node: x}]}} or {key: [x]}."""
    value = (node or {}).get(key)
    if isinstance(value, dict) and "edges" in value:
        return [e.get("node", {}) for e in value["edges"]]
    if isinstance(value, list):
        return value
    return []


def order_date(node):
    raw = node.get("processedAt") or node.get("createdAt") or ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return raw[:10]


def customer_name(node):
    addr = node.get("shippingAddress") or {}
    if addr.get("name"):
        return addr["name"]
    parts = [addr.get("firstName"), addr.get("lastName")]
    if not any(parts):
        cust = node.get("customer") or {}
        parts = [cust.get("firstName"), cust.get("lastName")]
    return " ".join(p for p in parts if p).strip()


def rows_for(node):
    """One row per line item; order-level fields repeat so the sheet sorts."""
    addr = node.get("shippingAddress") or {}
    cust = node.get("customer") or {}
    common = {
        "Order": node.get("name") or "",
        "Order Date": order_date(node),
        "Customer": customer_name(node),
        "Email": node.get("email") or cust.get("email") or "",
        "Phone": addr.get("phone") or node.get("phone") or cust.get("phone") or "",
        "Address 1": addr.get("address1") or "",
        "Address 2": addr.get("address2") or "",
        "City": addr.get("city") or "",
        "State/Province": addr.get("province") or addr.get("provinceCode") or "",
        "ZIP": addr.get("zip") or "",
        "Country": addr.get("country") or addr.get("countryCodeV2") or "",
        "Note": node.get("note") or "",
    }

    rows = []
    for item in unwrap(node, "lineItems"):
        image = item.get("image") or {}
        row = dict(common)
        row.update({
            "Product": item.get("title") or "",
            "Variant": (item.get("variantTitle") or "").replace("Default Title", ""),
            "SKU": item.get("sku") or "",
            "Qty": item.get("quantity") or 0,
            "Image Link": image.get("url") or "",
        })
        rows.append(row)

    if not rows:  # an order with no line items still deserves a visible row
        row = dict(common)
        row.update({"Product": "(no line items returned)", "Variant": "",
                    "SKU": "", "Qty": 0, "Image Link": ""})
        rows.append(row)
    return rows


def open_sheet(out_path):
    """Reuse today's file if a second run adds orders; else start fresh."""
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        return wb, wb.active, wb.active.max_row + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    for idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS[name]
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return wb, ws, 2


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--orders", required=True, help="JSON payload of order nodes")
    ap.add_argument("--out", required=True, help="destination .xlsx path")
    args = ap.parse_args()

    orders = load_orders(args.orders)
    if not orders:
        print("NO_ORDERS")  # routine treats this as a quiet no-op
        return

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb, ws, row_idx = open_sheet(args.out)

    # Stripe alternate orders so a multi-line order reads as one block.
    stripe = False
    for node in orders:
        for row in rows_for(node):
            for col_idx, name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[name])
                cell.alignment = Alignment(vertical="top", wrap_text=name in ("Address 1", "Note", "Product"))
                if stripe:
                    cell.fill = STRIPE_FILL
                if name == "Image Link" and row[name]:
                    cell.hyperlink = row[name]
                    cell.font = LINK_FONT
            row_idx += 1
        stripe = not stripe

    wb.save(args.out)
    print(f"WROTE {args.out} | orders={len(orders)} | rows={row_idx - 2}")


if __name__ == "__main__":
    main()
